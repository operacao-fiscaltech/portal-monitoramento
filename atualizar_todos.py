"""
atualizar_todos.py — Fiscaltech Monitoramento v2
=================================================
Roda todo dia às 07:00 (ou a cada hora) via Agendador de Tarefas.

O que faz:
  1. Varre PASTA_RAIZ buscando relatórios por contrato
  2. Extrai dados reais de todos os 12 relatórios (BIFF8 + XLSX)
  3. Gera dashboard_EPR.html (um por contrato)
  4. Atualiza portal.html com resumo de todos os contratos
  5. Publica tudo no GitHub via API REST

Relatórios suportados:
  - Aproveitamento_Geral_*.xls   (BIFF8)
  - Aproveitamento_Data_*.xls    (BIFF8)
  - Funcionamento_Fluxo_*.xlsx
  - Funcionamento_Infracoes_*.xlsx
  - Funcionamento_Teste_*.xlsx
  - Funcionamento_Horario_*.xls  (BIFF8)
  - LAP_Integral_24h_*.xls       (BIFF8)
  - LAP_Transicao_Manha_*.xls    (BIFF8)
  - LAP_Diurno_*.xls             (BIFF8)
  - LAP_Transicao_Tarde_*.xls    (BIFF8)
  - LAP_Noturno_*.xls            (BIFF8)
  - Lista_Geral_Pistas_*.xls     (BIFF8)

Dependencias: pip install openpyxl requests
"""

import os, re, struct, json, glob, base64, datetime, requests, csv, io
from openpyxl import load_workbook

# ============================================================
#  CONFIGURACAO
# ============================================================
PASTA_RAIZ     = r"C:\fiscaltech\relatorios"  # Pasta RAIZ — apenas ate aqui! Ex: C:\fiscaltech\relatorios
#                                    Nao inclua EPR, 2026, 03 etc.
#                                    O script localiza as subpastas automaticamente.
PASTA_LOCAL    = r"C:\fiscaltech\portal"
GITHUB_TOKEN   = "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"
GITHUB_USUARIO = "operacao-fiscaltech"
GITHUB_REPO    = "portal-monitoramento"
GITHUB_BRANCH  = "main"

LAP_OK    = 80
LAP_CRIT  = 60
APROV_OK   = 0.90   # Indice Aprov >= 0.90 = bom
APROV_CRIT = 0.70   # Indice Aprov < 0.70 = critico

CORES_CONTRATO = [
    "#0ea5e9","#22d3a0","#f59e0b","#8b5cf6",
    "#f43f5e","#fb923c","#38bdf8","#4ade80",
]

# ============================================================
#  PARSER OLE2/BIFF8 (XLS binario real)
# ============================================================

def _parse_ole2_workbook(filepath):
    """Extrai o stream Workbook de um arquivo OLE2/CFB."""
    with open(filepath, "rb") as f:
        raw = f.read()
    if raw[:8] != b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return None
    sector_size       = 2 ** struct.unpack_from('<H', raw, 30)[0]
    fat_sectors_count = struct.unpack_from('<I', raw, 44)[0]
    first_dir_sector  = struct.unpack_from('<I', raw, 48)[0]

    def sector_offset(sid):
        return 512 + sid * sector_size

    fat = {}
    for i in range(min(fat_sectors_count, 109)):
        sid = struct.unpack_from('<I', raw, 76 + i * 4)[0]
        if sid < 0xFFFFFFFB:
            off = sector_offset(sid)
            for j in range(sector_size // 4):
                fat[len(fat)] = struct.unpack_from('<I', raw, off + j * 4)[0]

    def follow_chain(start):
        chain, cur, seen = [], start, set()
        while cur < 0xFFFFFFFB and cur not in seen:
            chain.append(cur); seen.add(cur)
            cur = fat.get(cur, 0xFFFFFFFE)
        return chain

    def read_stream(chain):
        buf = b''
        for sec in chain:
            off = sector_offset(sec)
            buf += raw[off:off + sector_size]
        return buf

    dir_data = read_stream(follow_chain(first_dir_sector))
    for i in range(len(dir_data) // 128):
        off  = i * 128
        nlen = struct.unpack_from('<H', dir_data, off + 64)[0]
        name = dir_data[off:off + max(0, nlen - 2)].decode('utf-16-le', errors='ignore') if nlen > 2 else ''
        if name == 'Workbook':
            start = struct.unpack_from('<I', dir_data, off + 116)[0]
            size  = struct.unpack_from('<I', dir_data, off + 120)[0]
            return read_stream(follow_chain(start))[:size]
    return None


def _parse_biff8_cells(wb_data):
    """Extrai {(sheet_idx, row, col): value} e SST do stream BIFF8."""
    sst     = []
    cells   = {}
    sheet_i = -1

    i = 0
    while i < len(wb_data) - 4:
        rt = struct.unpack_from('<H', wb_data, i)[0]
        rl = struct.unpack_from('<H', wb_data, i + 2)[0]
        d  = wb_data[i + 4: i + 4 + rl]

        if rt == 0x0809 and len(d) >= 4:   # BOF
            if struct.unpack_from('<H', d, 2)[0] == 0x0010:
                sheet_i += 1

        elif rt == 0x00FC and len(d) >= 8:  # SST
            count = struct.unpack_from('<I', d, 4)[0]
            pos = 8
            for _ in range(min(count, 10000)):
                if pos + 3 > len(d): break
                slen  = struct.unpack_from('<H', d, pos)[0]
                flags = d[pos + 2] if pos + 2 < len(d) else 0
                pos  += 3
                if flags & 4:
                    rtc = struct.unpack_from('<H', d, pos)[0] if pos + 2 <= len(d) else 0
                    pos += 2 + rtc * 4
                if flags & 8:
                    sz  = struct.unpack_from('<I', d, pos)[0] if pos + 4 <= len(d) else 0
                    pos += 4 + sz
                if flags & 1:
                    end = pos + slen * 2
                    val = d[pos:end].decode('utf-16-le', errors='replace')
                else:
                    end = pos + slen
                    val = d[pos:end].decode('latin-1', errors='replace')
                pos = end
                sst.append(val)

        elif rt == 0x00FD and len(d) >= 10:  # LABELSST
            row = struct.unpack_from('<H', d, 0)[0]
            col = struct.unpack_from('<H', d, 2)[0]
            idx = struct.unpack_from('<I', d, 6)[0]
            cells[(sheet_i, row, col)] = sst[idx] if idx < len(sst) else ''

        elif rt == 0x0203 and len(d) >= 14:  # NUMBER
            row = struct.unpack_from('<H', d, 0)[0]
            col = struct.unpack_from('<H', d, 2)[0]
            cells[(sheet_i, row, col)] = struct.unpack_from('<d', d, 6)[0]

        elif rt == 0x027E and len(d) >= 10:  # RK
            row = struct.unpack_from('<H', d, 0)[0]
            col = struct.unpack_from('<H', d, 2)[0]
            rk  = struct.unpack_from('<I', d, 6)[0]
            if rk & 2:
                val = (rk >> 2) * (0.01 if rk & 1 else 1)
            else:
                val = struct.unpack_from('<d', bytes(4) + struct.pack('<I', rk & 0xFFFFFFFC))[0]
                if rk & 1: val /= 100
            cells[(sheet_i, row, col)] = val

        elif rt == 0x00BD and len(d) >= 6:   # MULRK
            row       = struct.unpack_from('<H', d, 0)[0]
            col_first = struct.unpack_from('<H', d, 2)[0]
            for k in range((len(d) - 6) // 6):
                off2 = 4 + k * 6
                rk   = struct.unpack_from('<I', d, off2 + 2)[0]
                if rk & 2:
                    val = (rk >> 2) * (0.01 if rk & 1 else 1)
                else:
                    val = struct.unpack_from('<d', bytes(4) + struct.pack('<I', rk & 0xFFFFFFFC))[0]
                    if rk & 1: val /= 100
                cells[(sheet_i, row, col_first + k)] = val

        elif rt == 0x0204 and len(d) >= 6:   # LABEL
            row  = struct.unpack_from('<H', d, 0)[0]
            col  = struct.unpack_from('<H', d, 2)[0]
            slen = struct.unpack_from('<H', d, 4)[0]
            cells[(sheet_i, row, col)] = d[6:6 + slen].decode('latin-1', errors='replace')

        i += 4 + rl

    return cells


def ler_xls_biff8(filepath):
    """Retorna dict {(row,col): value} para a sheet 0 de um XLS BIFF8."""
    wb_data = _parse_ole2_workbook(filepath)
    if not wb_data:
        return {}
    raw = _parse_biff8_cells(wb_data)
    return {(r, c): v for (si, r, c), v in raw.items() if si == 0}


def ler_xlsx(filepath):
    """Retorna dict {(row,col): value} para a sheet ativa de um XLSX (0-indexed)."""
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        cells = {}
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            for c_idx, val in enumerate(row):
                if val is not None:
                    cells[(r_idx, c_idx)] = val
        return cells
    except Exception as e:
        print(f"  [ERRO xlsx] {filepath}: {e}")
        return {}


# ============================================================
#  UTILITARIOS
# ============================================================

def encontrar_arquivo(pasta, *padroes):
    """Busca arquivos por padrão glob (case-insensitive no fallback)."""
    if not pasta or not os.path.isdir(pasta):
        return None
    for padrao in padroes:
        # Glob direto (funciona no Windows que já é case-insensitive)
        matches = sorted(glob.glob(os.path.join(pasta, padrao)), reverse=True)
        if matches:
            return matches[0]
        # Fallback: busca manual para Linux/testes
        # Converte "Arquivo_*.xls" -> regex "^Arquivo_.*\.xls$"
        partes = padrao.split('*')
        partes_esc = [re.escape(p) for p in partes]
        padrao_re = re.compile('^' + '.*'.join(partes_esc) + '$', re.IGNORECASE)
        hits = sorted(
            [f for f in os.listdir(pasta) if padrao_re.match(f)],
            reverse=True
        )
        if hits:
            return os.path.join(pasta, hits[0])
    return None


def cel_float(cells, row, col, default=0.0):
    v = cells.get((row, col), default)
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def cel_str(cells, row, col, default=''):
    v = cells.get((row, col), default)
    return str(v).strip() if v is not None else default


def excel_date_to_str(serial):
    """Converte serial Excel (dias desde 1900-01-01) para ISO."""
    try:
        base = datetime.date(1899, 12, 30)
        return (base + datetime.timedelta(days=int(serial))).isoformat()
    except Exception:
        return ''


# ============================================================
#  LEITURA: LISTA GERAL DE PISTAS
# ============================================================

def ler_lista_pistas(pasta):
    path = encontrar_arquivo(pasta,
        "Lista_Geral_Pistas_*.xls", "Lista_Geral_Pistas_*.xlsx",
        "ListaGeralPistaServlet*.xls", "ListaGeralPistaServlet*.xlsx")
    if not path:
        arquivos = os.listdir(pasta) if os.path.isdir(pasta) else []
        xls_na_pasta = [f for f in arquivos if f.lower().endswith(('.xls','.xlsx'))]
        print(f"  [AVISO] Lista_Geral_Pistas nao encontrada em: {pasta}")
        if xls_na_pasta:
            print(f"    Arquivos XLS encontrados: {xls_na_pasta[:5]}")
        else:
            print(f"    Nenhum .xls/.xlsx na pasta. Conteudo: {arquivos[:8]}")
        return []

    cells = ler_xls_biff8(path) if path.endswith('.xls') else ler_xlsx(path)
    equipamentos = {}  # serie_int -> dict

    row = 4  # primeira linha de dados (0-indexed)
    while True:
        item = cel_str(cells, row, 1)
        if not item:
            if row > 200: break
            row += 1
            continue

        serie_raw = cel_str(cells, row, 2)   # ex: "FSC110-0546"
        m = re.search(r'(\d{3,4})$', serie_raw)
        if not m:
            row += 1
            continue
        serie = int(m.group(1))

        local     = cel_str(cells, row, 3)
        sentido   = cel_str(cells, row, 5)
        faixa     = cel_str(cells, row, 6)
        cod_cli   = cel_str(cells, row, 7)
        vel       = int(cel_float(cells, row, 13, 80))
        dt_afer   = cel_float(cells, row, 16, 0)
        venc_afer = cel_float(cells, row, 17, 0)

        if serie not in equipamentos:
            equipamentos[serie] = {
                "serie":      str(serie),
                "local":      local,
                "cod":        cod_cli,
                "sentido":    sentido,
                "vel":        vel,
                "ult_afer":   excel_date_to_str(dt_afer)   if dt_afer   > 1 else '',
                "venc_afer":  excel_date_to_str(venc_afer) if venc_afer > 1 else '',
                "faixas":     [],
            }

        faixa_id = str(int(float(faixa))) if re.match(r'^\d+\.?\d*$', faixa) else faixa
        faixas_existentes = [f["id"] for f in equipamentos[serie]["faixas"]]
        if faixa_id not in faixas_existentes:
            equipamentos[serie]["faixas"].append({"id": faixa_id})

        row += 1

    result = []
    for serie in sorted(equipamentos):
        eq = equipamentos[serie]
        eq["faixas"] = sorted(eq["faixas"], key=lambda f: f["id"])
        result.append(eq)

    print(f"  [Pistas] {len(result)} equipamentos, {sum(len(e['faixas']) for e in result)} faixas")
    return result


# ============================================================
#  LEITURA: FUNCIONAMENTO (Fluxo / Infracoes / Teste)
# ============================================================

def _extrair_cabecalho_dias(cells, row_header=6, col_inicio=10, col_fim=42):
    dias = []
    for c in range(col_inicio, col_fim):
        v = cel_str(cells, row_header, c)
        if re.match(r'^\d{1,2}\s+\S+', v):
            dias.append((c, v.strip()))
    return dias


def ler_funcionamento(path):
    """Retorna {(serie_int, faixa_str): {label_dia: valor}} e lista de labels."""
    cells    = ler_xlsx(path) if path.endswith('.xlsx') else ler_xls_biff8(path)
    dias_cols = _extrair_cabecalho_dias(cells)
    dados    = {}

    row = 7
    while True:
        item = cells.get((row, 1))
        if item is None:
            if row > 300: break
            row += 1
            continue
        if isinstance(item, str) and 'equip' in item.lower():
            break

        serie_raw = cel_float(cells, row, 2, 0)
        if serie_raw == 0:
            row += 1
            continue
        serie = int(serie_raw)
        faixa = cel_str(cells, row, 7).strip() or '1'

        linha_dias = {label: cel_float(cells, row, col) for col, label in dias_cols}
        dados[(serie, faixa)] = linha_dias
        row += 1

    return dados, [label for _, label in dias_cols]


# ============================================================
#  LEITURA: APROVEITAMENTO GERAL
# ============================================================
#  Colunas (0-indexed após row 7):
#  1=Item, 2=Série, 3=Local, 4=vazio, 5=Faixa, 6=Data(vazio),
#  7=Tráfego, 8=Vel.Média, 9=Total Infrac, 10=Inc.Técnica,
#  11=Inc.Não Técnica, 12=Total Incons, 13=Total Consist, 14=Em Proces,
#  15=Indice Aprov, 16=% Aprov
# ============================================================

def ler_aproveitamento_geral(path):
    """Retorna {(serie_int, faixa_str): {campos...}}"""
    cells = ler_xls_biff8(path) if path.endswith('.xls') else ler_xlsx(path)
    dados = {}
    row = 8
    while True:
        item = cells.get((row, 1))
        if item is None:
            if row > 500: break
            row += 1
            continue
        serie_raw = cel_float(cells, row, 2, 0)
        if serie_raw == 0:
            row += 1
            continue
        serie = int(serie_raw)
        faixa = str(int(cel_float(cells, row, 5, 1)))

        dados[(serie, faixa)] = {
            "trafego":       int(cel_float(cells, row, 7)),
            "vel_media":     round(cel_float(cells, row, 8), 1),
            "total_infra":   int(cel_float(cells, row, 9)),
            "inc_tec":       int(cel_float(cells, row, 10)),
            "inc_n_tec":     int(cel_float(cells, row, 11)),
            "total_incons":  int(cel_float(cells, row, 12)),
            "total_consist": int(cel_float(cells, row, 13)),
            "em_proc":       int(cel_float(cells, row, 14)),
            "indice_aprov":  round(cel_float(cells, row, 15), 4),
            "pct_aprov":     round(cel_float(cells, row, 16), 4),
        }
        row += 1

    print(f"  [Aprov.Geral] {len(dados)} faixas")
    return dados


# ============================================================
#  LEITURA: APROVEITAMENTO POR DATA
# ============================================================

def ler_aproveitamento_data(path):
    """Retorna {(serie, faixa, data_iso): {campos...}}"""
    cells = ler_xls_biff8(path) if path.endswith('.xls') else ler_xlsx(path)
    dados = {}
    row = 8
    while True:
        item = cells.get((row, 1))
        if item is None:
            if row > 10000: break
            row += 1
            continue
        serie_raw = cel_float(cells, row, 2, 0)
        if serie_raw == 0:
            row += 1
            continue
        serie = int(serie_raw)
        faixa = str(int(cel_float(cells, row, 5, 1)))
        data  = cel_str(cells, row, 6)  # 'YYYY-MM-DD'
        if not data:
            row += 1
            continue
        dados[(serie, faixa, data)] = {
            "trafego":      int(cel_float(cells, row, 7)),
            "vel_media":    round(cel_float(cells, row, 8), 1),
            "total_infra":  int(cel_float(cells, row, 9)),
            "inc_tec":      int(cel_float(cells, row, 10)),
            "inc_n_tec":    int(cel_float(cells, row, 11)),
            "indice_aprov": round(cel_float(cells, row, 15), 4),
            "pct_aprov":    round(cel_float(cells, row, 16), 4),
        }
        row += 1

    print(f"  [Aprov.Data] {len(dados)} registros")
    return dados


# ============================================================
#  LEITURA: LAP (todos os subperiodos)
# ============================================================

def ler_lap(path):
    """Retorna {(serie_int, faixa_str): {label_dia: pct_int|None}} e lista de labels."""
    cells = ler_xls_biff8(path) if path.endswith('.xls') else ler_xlsx(path)

    dias_cols = []
    for c in range(8, 50):
        v = cel_str(cells, 10, c)
        if re.match(r'^\d{1,2}\s+\S+', v):
            dias_cols.append((c, v.strip()))

    dados = {}
    row = 11
    while True:
        item = cells.get((row, 1))
        if item is None:
            if row > 200: break
            row += 1
            continue
        serie_raw = cel_float(cells, row, 2, 0)
        if serie_raw == 0:
            row += 1
            continue
        serie = int(serie_raw)
        faixa = str(int(cel_float(cells, row, 4, 1)))

        linha_dias = {}
        for col, label in dias_cols:
            v = cel_float(cells, row, col, -1)
            linha_dias[label] = int(v) if v >= 0 else None

        dados[(serie, faixa)] = linha_dias
        row += 1

    return dados, [label for _, label in dias_cols]


# ============================================================
#  LEITURA: FUNCIONAMENTO HORARIO
# ============================================================

def ler_funcionamento_horario(path):
    """Retorna {(serie, faixa, data_iso): [v0h..v23h]}"""
    cells = ler_xls_biff8(path) if path.endswith('.xls') else ler_xlsx(path)
    dados = {}
    row = 5
    while True:
        serie_raw = cel_float(cells, row, 1, 0)
        if serie_raw == 0:
            if row > 5000: break
            row += 1
            continue
        serie  = int(serie_raw)
        faixa  = cel_str(cells, row, 7).strip() or '1'
        data_s = cel_float(cells, row, 8, 0)
        data   = excel_date_to_str(data_s) if data_s > 1 else ''
        if not data or not faixa:
            row += 1
            continue
        horas = [int(cel_float(cells, row, 9 + h, 0)) for h in range(24)]
        dados[(serie, faixa, data)] = horas
        row += 1

    print(f"  [Horario] {len(dados)} registros")
    return dados


# ============================================================
#  MONTAGEM DOS DADOS DO CONTRATO
# ============================================================

def montar_dados_contrato(pasta, id_contrato, ano, mes):
    print(f"\n[{id_contrato}] Lendo relatórios de: {pasta}")

    equipamentos = ler_lista_pistas(pasta)
    if not equipamentos:
        return None

    # Funcionamento
    pf = encontrar_arquivo(pasta, "Funcionamento_Fluxo_*.xlsx",     "Funcionamento_Fluxo_*.xls")
    pi = encontrar_arquivo(pasta, "Funcionamento_Infracoes_*.xlsx",  "Funcionamento_Infracoes_*.xls")
    pt = encontrar_arquivo(pasta, "Funcionamento_Teste_*.xlsx",      "Funcionamento_Teste_*.xls")
    ph = encontrar_arquivo(pasta, "Funcionamento_Horario_*.xls",     "Funcionamento_Horario_*.xlsx")

    d_fluxo, dias_labels = ler_funcionamento(pf) if pf else ({}, [])
    d_infra, _           = ler_funcionamento(pi) if pi else ({}, [])
    d_teste, _           = ler_funcionamento(pt) if pt else ({}, [])
    d_horario            = ler_funcionamento_horario(ph) if ph else {}

    # LAP
    d_lap = {}
    lap_dias = []
    for nome, padrao in [
        ("integral",  "LAP_Integral_24h_*.xls"),
        ("tr_manha",  "LAP_Transicao_Manha_*.xls"),
        ("diurno",    "LAP_Diurno_*.xls"),
        ("tr_tarde",  "LAP_Transicao_Tarde_*.xls"),
        ("noturno",   "LAP_Noturno_*.xls"),
    ]:
        p = encontrar_arquivo(pasta, padrao)
        if p:
            dd, lds = ler_lap(p)
            d_lap[nome] = dd
            if not lap_dias: lap_dias = lds
            print(f"  LAP {nome}: {len(dd)} faixas")
        else:
            d_lap[nome] = {}

    # Aproveitamento
    pa_g = encontrar_arquivo(pasta, "Aproveitamento_Geral_*.xls")
    pa_d = encontrar_arquivo(pasta, "Aproveitamento_Data_*.xls")
    d_aprov_geral = ler_aproveitamento_geral(pa_g) if pa_g else {}
    d_aprov_data  = ler_aproveitamento_data(pa_d)  if pa_d else {}

    # Montar dados por faixa
    for eq in equipamentos:
        serie = int(eq["serie"])
        for fx in eq["faixas"]:
            faixa = fx["id"]
            key   = (serie, faixa)

            fx["fluxo"] = d_fluxo.get(key, {})
            fx["infra"] = d_infra.get(key, {})
            fx["teste"] = d_teste.get(key, {})

            fx["lap"] = {
                "integral": d_lap.get("integral", {}).get(key, {}),
                "tr_manha": d_lap.get("tr_manha", {}).get(key, {}),
                "diurno":   d_lap.get("diurno",   {}).get(key, {}),
                "tr_tarde": d_lap.get("tr_tarde", {}).get(key, {}),
                "noturno":  d_lap.get("noturno",  {}).get(key, {}),
            }

            aprov = d_aprov_geral.get(key)
            fx["aproveitamento"] = aprov or {
                "trafego": 0, "vel_media": 0, "total_infra": 0,
                "inc_tec": 0, "inc_n_tec": 0, "total_incons": 0,
                "total_consist": 0, "em_proc": 0, "indice_aprov": 0, "pct_aprov": 0
            }

            fx["aprov_por_data"] = {
                data: vals
                for (s, f, data), vals in d_aprov_data.items()
                if s == serie and f == faixa
            }

            horario_key = max(
                [(s, f2, d) for (s, f2, d) in d_horario if s == serie and f2 == faixa],
                default=None
            )
            fx["horario"] = d_horario.get(horario_key, [0]*24)

    return {
        "id":           id_contrato,
        "nome":         id_contrato,
        "dias":         dias_labels or lap_dias,
        "faixas": equipamentos,
    }


# ============================================================
#  RESUMO PARA O PORTAL
# ============================================================

def dias_rest(venc):
    if not venc: return 999
    try:
        return (datetime.date.fromisoformat(venc) - datetime.date.today()).days
    except Exception:
        return 999


def detectar_rodovias(equipamentos):
    rv = set()
    for eq in equipamentos:
        for m in re.findall(r'(?:BR|PR|SP|SC)-\d+', eq.get("local", ""), re.I):
            rv.add(m.upper())
    return sorted(rv)[:5]


def calcular_resumo(dados, cor, arquivo_html):
    eqs  = dados["faixas"]
    dias = dados["dias"]
    crit = warn = ok = off = 0
    lap_vals = []
    aprov_vals = []
    fluxo_spark = []
    faixas_sample = []

    for d_label in dias:
        total = sum(
            fx.get("fluxo", {}).get(d_label, 0)
            for eq in eqs for fx in eq["faixas"]
        )
        fluxo_spark.append(int(total))

    for eq in eqs:
        for fx in eq["faixas"]:
            lap_dict   = fx.get("lap", {}).get("integral", {})
            lap_ultimo = next((v for v in reversed(list(lap_dict.values())) if v is not None), None)
            aprov      = fx.get("aproveitamento", {})
            indice     = aprov.get("indice_aprov", 0) or 0
            flxh       = fx.get("fluxo", {}).get(dias[-1], 0) if dias else 0
            dias_teste = sum(1 for v in fx.get("teste", {}).values() if v and v > 0)

            if lap_ultimo is not None and lap_ultimo >= 0:
                lap_vals.append(lap_ultimo)
            if indice > 0:
                aprov_vals.append(indice)

            is_off  = (flxh == 0 and dias_teste == 0)
            is_crit = not is_off and (
                (indice > 0 and indice < APROV_CRIT) or
                (lap_ultimo is not None and lap_ultimo < LAP_CRIT)
            )
            is_warn = not is_off and not is_crit and (
                (indice >= APROV_CRIT and indice < APROV_OK) or
                (lap_ultimo is not None and lap_ultimo >= LAP_CRIT and lap_ultimo < LAP_OK)
            )

            if is_off:    off  += 1
            elif is_crit: crit += 1
            elif is_warn: warn += 1
            else:         ok   += 1

            st = "off" if is_off else ("crit" if is_crit else ("warn" if is_warn else "ok"))
            faixas_sample.append({"serie": eq["serie"], "id": fx["id"], "status": st})

    return {
        "id":              dados["id"],
        "nome":            dados["nome"],
        "arquivo":         arquivo_html,
        "rodovias":        detectar_rodovias(eqs),
        "uf":              "PR",
        "totalEquip":      len(eqs),
        "totalFaixas":     sum(len(eq["faixas"]) for eq in eqs),
        "fluxoHoje":       fluxo_spark[-1] if fluxo_spark else 0,
        "lapMedio":        round(sum(lap_vals) / len(lap_vals), 1) if lap_vals else 0,
        "aprovMedio":      round(sum(aprov_vals) / len(aprov_vals), 3) if aprov_vals else 0,
        "faixasCrit":      crit,
        "faixasWarn":      warn,
        "faixasOk":        ok,
        "faixasOff":       off,
        "afericoesVenc60": sum(1 for eq in eqs if dias_rest(eq.get("venc_afer", "")) <= 60),
        "faixasSample":    faixas_sample[:16],
        "fluxoSpark":      fluxo_spark,
        "cor":             cor,
        "atualizado":      datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
    }


# ============================================================

# ============================================================
#  LEITURA: CONSULTA INFRAÇÕES (CSV)
# ============================================================

def ler_consulta_infracoes(pasta):
    """
    Le consulta_infracoes_YYYYMMDD.csv gerado pelo baixar_relatorios.py.
    Retorna (top_js, sp_js, enq_js) prontas para injetar no dashboard,
    ou (None, None, None) se nao encontrar o CSV.
    """
    from collections import defaultdict

    path = encontrar_arquivo(pasta,
        "consulta_infracoes_*.csv",
        "consulta_infrações_*.csv",
        "consulta_infracoes*.csv",
        "consulta_infra*.csv")
    if not path:
        print("  [AVISO] consulta_infracoes CSV nao encontrado em: " + str(pasta))
        return None, None, None

    print(f"  [CSV] {os.path.basename(path)}")

    for enc in ('utf-8-sig', 'latin-1', 'utf-8', 'cp1252'):
        try:
            with open(path, 'r', encoding=enc) as f:
                sample = f.read(512)
            if any(k in sample for k in ('Cod. Pista', 'Razao Tecnica', 'Descricao')):
                encoding = enc
                break
        except Exception:
            pass
    else:
        encoding = 'latin-1'

    with open(path, 'r', encoding=encoding) as f:
        reader = csv.DictReader(f, delimiter=';')
        rows = list(reader)

    print(f"  [CSV] {len(rows)} registros (encoding: {encoding})")

    by_sp   = defaultdict(lambda: {
        'tec': 0, 'ntec_real': 0, 'consist': 0, 'total': 0,
        'tec_top': defaultdict(int), 'ntec_top': defaultdict(int),
        'nome': '', 'sentido': ''})
    by_spd  = defaultdict(lambda: {
        'tec': 0, 'ntec': 0, 'ntec_real': 0, 'consist': 0, 'total': 0,
        'by_date': defaultdict(lambda: {
            'tec': 0, 'ntec': 0, 'ntec_real': 0, 'consist': 0, 'total': 0})})
    enq_map = {'74550': 'Velocidade', '74630': 'Vel.+Peso', '74710': 'Outros'}
    enq_dia  = defaultdict(lambda: defaultdict(int))
    enq_ser  = defaultdict(lambda: defaultdict(int))

    # Mapa série_numerica → cod_cliente (ex: "546" → "EPR007")
    serie_to_cod = {}

    for r in rows:
        local = r.get('Local', '').strip()          # série numérica: "546"
        cc    = r.get('Cod. Pista Cliente', '').strip()  # "EPR007" (pode estar vazio)
        pi    = r.get('Pista', '').strip()
        desc  = r.get('Descricao', '').strip()
        rt    = r.get('Razao Tecnica', '').strip()
        nm    = r.get('Nome Pista', '').strip()
        se    = r.get('Sentido', '').strip()
        enq   = r.get('Enquadramento', '').strip()
        dt    = r.get('Data Imagem', '').strip()
        dia   = dt.split('/')[0] if '/' in dt else '??'
        # Chave primária: Local|Pista (série numérica, sempre preenchida)
        key   = f"{local}|{pi}"
        # Registrar mapeamento serie → cod_cliente se disponível
        if cc and cc not in ('0', '') and local:
            serie_to_cod[local] = cc

        by_sp[key]['total'] += 1
        by_sp[key]['nome']   = nm
        by_sp[key]['sentido']= se
        by_spd[key]['total'] += 1
        by_spd[key]['by_date'][dia]['total'] += 1

        if rt == 'Técnica':
            by_sp[key]['tec'] += 1
            by_sp[key]['tec_top'][desc] += 1
            by_spd[key]['tec'] += 1
            by_spd[key]['by_date'][dia]['tec'] += 1
        else:
            by_spd[key]['ntec'] += 1
            by_spd[key]['by_date'][dia]['ntec'] += 1
            if desc == 'Consistente':
                by_sp[key]['consist'] += 1
                by_spd[key]['consist'] += 1
                by_spd[key]['by_date'][dia]['consist'] += 1
            else:
                by_sp[key]['ntec_real'] += 1
                by_sp[key]['ntec_top'][desc] += 1
                by_spd[key]['ntec_real'] += 1
                by_spd[key]['by_date'][dia]['ntec_real'] += 1

        enq_label = enq_map.get(enq, 'Outros')
        enq_dia[enq_label][dia] += 1
        enq_ser[key][enq_label] += 1

    # Gerar top_data com chave numérica (546|1) E alias EPR007|1
    top_data = {}
    for k, v in by_sp.items():
        serie, faixa = k.split('|', 1)
        entry = {
            'nome': v['nome'], 'sentido': v['sentido'],
            'total': v['total'], 'tec': v['tec'],
            'ntec_real': v['ntec_real'], 'consist': v['consist'],
            'tec_top':  dict(sorted(v['tec_top'].items(),  key=lambda x: -x[1])[:5]),
            'ntec_top': dict(sorted(v['ntec_top'].items(), key=lambda x: -x[1])[:5]),
        }
        top_data[k] = entry  # chave numerica: "546|1"
        cod = serie_to_cod.get(serie)
        if cod:
            top_data[f"{cod}|{faixa}"] = entry  # alias EPR: "EPR007|1"

    sp_clean = {k: {
        'tec': v['tec'], 'ntec': v['ntec'], 'ntec_real': v['ntec_real'],
        'consist': v['consist'], 'total': v['total'],
        'by_date': {d: dict(dv) for d, dv in v['by_date'].items()}
    } for k, v in by_spd.items()}

    enq_dia_c = {e: dict(d) for e, d in enq_dia.items()}
    enq_ser_c = {k: dict(v) for k, v in enq_ser.items()}

    print(f"  [CSV] {len(top_data)} faixas processadas")
    return (
        f"const TOP_INFRA = {json.dumps(top_data, ensure_ascii=False)};",
        f"const INFRA_BY_SP = {json.dumps(sp_clean, ensure_ascii=False)};",
        f"const INFRA_ENQ_DIA = {json.dumps(enq_dia_c, ensure_ascii=False)};\n"
        f"const INFRA_ENQ_SERIE = {json.dumps(enq_ser_c, ensure_ascii=False)};"
    )

#  GERACAO HTML
# ============================================================

def gerar_dashboard(template_path, dados, id_contrato, csv_js=None):
    with open(template_path, "r", encoding="utf-8") as f:
        html = f.read()

    # Diagnóstico — confirmar que o template está limpo
    db_antes = re.search(r'const DB = \[(.*?)\];', html, flags=re.DOTALL)
    if db_antes:
        conteudo = db_antes.group(1).strip()
        if conteudo:
            print(f"  [AVISO] Template ainda tem dados ({len(conteudo)} chars) — forçando limpeza")
            html = re.sub(r'const DB = \[.*?\];', 'const DB = [];', html, flags=re.DOTALL)
        else:
            print(f"  [OK] Template limpo, injetando dados novos")
    else:
        print(f"  [ERRO] 'const DB' não encontrado no template!")

    db_json = json.dumps([dados], ensure_ascii=False, separators=(',', ':'))
    html = re.sub(r'const DB = \[\];', f'const DB = {db_json};', html)

    # Timestamp de geração — garante que o conteúdo sempre muda
    agora = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    html = re.sub(r'<!-- GERADO:.*?-->', '', html)
    html = html.replace('</head>', f'<!-- GERADO: {agora} -->\n</head>', 1)

    # Injetar dados do CSV de infrações se disponível
    if csv_js:
        top_js, sp_js, enq_js = csv_js
        if top_js:
            html = re.sub(r'const TOP_INFRA = \{.*?\};', top_js, html, flags=re.DOTALL)
        if sp_js:
            html = re.sub(r'const INFRA_BY_SP = \{.*?\};', sp_js, html, flags=re.DOTALL)
        if enq_js:
            html = re.sub(r'const INFRA_ENQ_DIA = .*?;\n', '', html, flags=re.DOTALL)
            html = re.sub(r'const INFRA_ENQ_SERIE = \{.*?\};', '', html, flags=re.DOTALL)
            html = html.replace(sp_js, sp_js + '\n' + enq_js, 1)

    print(f"  [GERADO] {agora} — {len(db_json)} bytes de dados")
    return html



def gerar_portal(template_path, resumos):
    with open(template_path, "r", encoding="utf-8") as f:
        html = f.read()
    html = re.sub(
        r'const CONTRATOS = \[.*?\];',
        f'const CONTRATOS = {json.dumps(resumos, ensure_ascii=False, indent=2)};',
        html, flags=re.DOTALL
    )
    return html


# ============================================================
#  GITHUB
# ============================================================

def gh_headers():
    return {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}


def gh_sha(caminho):
    url = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{caminho}"
    r = requests.get(url, headers=gh_headers(), params={"ref": GITHUB_BRANCH})
    if r.status_code == 200:
        return r.json().get("sha")
    if r.status_code == 401:
        print(f"  [GITHUB] Autenticacao falhou — verifique GITHUB_TOKEN")
    elif r.status_code == 404:
        pass  # arquivo ainda nao existe, normal na primeira vez
    return None


def gh_upload(caminho_local, caminho_repo):
    with open(caminho_local, "rb") as f:
        conteudo = base64.b64encode(f.read()).decode()
    sha = gh_sha(caminho_repo)
    url  = f"https://api.github.com/repos/{GITHUB_USUARIO}/{GITHUB_REPO}/contents/{caminho_repo}"
    data = {
        "message": f"Auto {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "content": conteudo, "branch": GITHUB_BRANCH
    }
    if sha:
        data["sha"] = sha
    r = requests.put(url, headers=gh_headers(), json=data)
    ok = r.status_code in (200, 201)
    if ok:
        print(f"  OK ({r.status_code}) — {caminho_repo}")
    else:
        msg = r.json().get("message", r.text[:120]) if r.content else r.status_code
        print(f"  ERRO {r.status_code} — {caminho_repo}")
        print(f"    Motivo: {msg}")
        print(f"    URL: {url}")
        print(f"    Usuario: {GITHUB_USUARIO}  Repo: {GITHUB_REPO}  Branch: {GITHUB_BRANCH}")
    return ok


# ============================================================
#  MAIN
# ============================================================

def pasta_mes(raiz, cid, ano=None, mes=None):
    """
    Localiza a pasta de relatorios do contrato.
    Suporta: relatorios/EPR/2026/03, relatorios/EPR/2026/3,
    relatorios/EPR/marco, relatorios/EPR/03, relatorios/EPR (flat).
    Retorna a primeira que existir e tiver .xls/.xlsx dentro.
    """
    if ano is None or mes is None:
        hoje = datetime.date.today()
        ano  = hoje.year
        mes  = hoje.month

    nomes_mes = {
        1:"janeiro",2:"fevereiro",3:"marco",4:"abril",
        5:"maio",6:"junho",7:"julho",8:"agosto",
        9:"setembro",10:"outubro",11:"novembro",12:"dezembro",
    }

    base = os.path.join(raiz, cid)
    candidatas = [
        os.path.join(base, str(ano), f"{mes:02d}"),
        os.path.join(base, str(ano), str(mes)),
        os.path.join(base, str(ano), nomes_mes[mes]),
        os.path.join(base, f"{mes:02d}"),
        os.path.join(base, str(mes)),
        os.path.join(base, nomes_mes[mes]),
        base,
    ]

    def tem_relatorio(p):
        if not os.path.isdir(p):
            return False
        return any(
            f.lower().endswith(('.xls', '.xlsx'))
            for f in os.listdir(p)
        )

    for c in candidatas:
        if tem_relatorio(c):
            return c

    # Fallback: busca recursiva dentro de base procurando o mês correto
    for dirpath, dirnames, filenames in os.walk(base):
        if any(f.lower().endswith(('.xls', '.xlsx')) for f in filenames):
            # Verifica se o caminho contém referência ao mês/ano atual
            partes = dirpath.replace('\\', '/').lower().split('/')
            ref_mes = [str(mes), f"{mes:02d}", nomes_mes[mes]]
            ref_ano = [str(ano)]
            if any(p in ref_mes or p in ref_ano for p in partes[-3:]):
                return dirpath

    # Último recurso: retorna a pasta base mesmo sem arquivos
    return base


def _e_contrato(nome):
    """
    Retorna True se o nome de pasta parece ser um ID de contrato
    (ex: EPR, ABC, DNIT, ARTESP) e não um ano (2025, 2026) ou mês (01..12).
    """
    if re.match(r'^\d{4}$', nome):   # ano: 2024, 2025, 2026...
        return False
    if re.match(r'^\d{1,2}$', nome): # mês: 1..12
        return False
    return True


def main():
    hoje = datetime.date.today()
    print(f"\n{'='*55}\n  Fiscaltech — {hoje.strftime('%d/%m/%Y %H:%M')}\n{'='*55}\n")

    if not os.path.exists(PASTA_RAIZ):
        print(f"ERRO: pasta nao encontrada: {PASTA_RAIZ}"); return

    # Filtra subpastas que sejam realmente IDs de contrato (não anos nem meses)
    contratos = [d for d in sorted(os.listdir(PASTA_RAIZ))
                 if os.path.isdir(os.path.join(PASTA_RAIZ, d))
                 and not d.startswith('.')
                 and _e_contrato(d)]
    if not contratos:
        print("ERRO: nenhuma subpasta de contrato encontrada"); return
    print(f"Contratos: {contratos}\n")

    os.makedirs(PASTA_LOCAL, exist_ok=True)
    template_dash = os.path.join(PASTA_LOCAL, "dashboard_EPR_template.html")
    # Fallback: se não existir template, usa dashboard_EPR.html mas avisa
    if not os.path.exists(template_dash):
        fallback = os.path.join(PASTA_LOCAL, "dashboard_EPR.html")
        if os.path.exists(fallback):
            print(f"  [AVISO] dashboard_EPR_template.html não encontrado.")
            print(f"  Renomeando dashboard_EPR.html → dashboard_EPR_template.html")
            import shutil
            shutil.copy(fallback, template_dash)
        else:
            print(f"  ERRO: template do dashboard não encontrado em {PASTA_LOCAL}")
            return
    template_port = os.path.join(PASTA_LOCAL, "portal.html")

    resumos = []
    gerados = []

    for i, cid in enumerate(contratos):
        try:
            p = pasta_mes(PASTA_RAIZ, cid, hoje.year, hoje.month)
            print(f"  [{cid}] Lendo relatorios de: {p}")
            dados = montar_dados_contrato(p, cid, hoje.year, hoje.month)
            if not dados:
                continue
            nome_h = f"dashboard_{cid}.html"
            cor    = CORES_CONTRATO[i % len(CORES_CONTRATO)]
            resumo = calcular_resumo(dados, cor, nome_h)
            resumos.append(resumo)

            csv_js = ler_consulta_infracoes(p)
            html = gerar_dashboard(template_dash, dados, cid, csv_js=csv_js)
            caminho = os.path.join(PASTA_LOCAL, nome_h)
            with open(caminho, "w", encoding="utf-8") as f:
                f.write(html)
            gerados.append((caminho, nome_h))
            print(f"  OK — {nome_h}\n")
        except Exception as e:
            import traceback
            print(f"  ERRO [{cid}]: {e}")
            traceback.print_exc()

    if resumos:
        html_p = gerar_portal(template_port, resumos)
        cp = os.path.join(PASTA_LOCAL, "portal.html")
        with open(cp, "w", encoding="utf-8") as f:
            f.write(html_p)
        gerados.append((cp, "portal.html"))
        print(f"portal.html atualizado ({len(resumos)} contratos)\n")

    # Publicar configurador no GitHub
    cfg_local = os.path.join(PASTA_LOCAL, "configurar_fiscaltech.html")
    if os.path.exists(cfg_local):
        gerados.append((cfg_local, "configurar_fiscaltech.html"))
    else:
        print(f"  [AVISO] configurar_fiscaltech.html não encontrado em {PASTA_LOCAL}")

    print(f"[GITHUB] Publicando {len(gerados)} arquivo(s)...")
    ok = sum(1 for c, r in gerados if gh_upload(c, r))
    print(f"\nConcluido: {ok}/{len(gerados)} publicados")
    print(f"Portal: https://{GITHUB_USUARIO}.github.io/{GITHUB_REPO}/portal.html\n")


if __name__ == "__main__":
    main()
